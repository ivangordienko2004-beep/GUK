from __future__ import annotations

import json
from functools import lru_cache
from pathlib import Path
import uuid

import pandas as pd
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MAPPING_DIR = Path(settings.BASE_DIR) / 'Mapping'
OFFICER_VUS_PATH = MAPPING_DIR / 'officer_vus.json'
NONOFFICE_VUS_PATH = MAPPING_DIR / 'nonoffice_vus.json'
NONOFFICE_POSITION_PATH = MAPPING_DIR / 'nonoffice_position.json'


REQUIRED_COLUMNS = [
    'okrug_vuza',
    'ovu_otv_podgotovku',
    'nazvanie_vuza',
    'vus_no',
    'vus_naimenovanie',
    'doljnost_no',
    'doljnost_naimenovanie',
    'sbor_stazhirovka',
    'programma_podgotovki',
    'mesto_provedeniya_uchebnogo_sbora',
    'planiruetsya_prepodavatelej',
    'planiruetsya_studentov',
    'srok_provedeniya_nachalo',
    'srok_provedeniya_okonchanie',
    'fio_otvetstvennogo',
    'mobilnyy',
]


OUTPUT_COLUMNS = [
    ('', 'ОКРУГ ВУЗа', '1', 'okrug_vuza'),
    ('', 'ОВУ, отв. за подготовку', '2', 'ovu_otv_podgotovku'),
    ('', 'Наименование ВУЗа', '3', 'nazvanie_vuza'),
    ('ВУС', '№', '4', 'vus_no'),
    ('ВУС', 'Наименование', '5', 'vus_naimenovanie'),
    ('Должность', '№', '6', 'doljnost_no'),
    ('Должность', 'Наименование', '7', 'doljnost_naimenovanie'),
    ('', 'Сбор/стаж', '8', 'sbor_stazhirovka'),
    ('', 'Программа', '9', 'programma_podgotovki'),
    ('', 'Место проведения', '10', 'mesto_provedeniya_uchebnogo_sbora'),
    ('Планируется', 'преподавателей', '11', 'planiruetsya_prepodavatelej'),
    ('Планируется', 'студентов', '12', 'planiruetsya_studentov'),
    ('Сроки проведения', 'начало', '13', 'srok_provedeniya_nachalo'),
    ('Сроки проведения', 'окончание', '14', 'srok_provedeniya_okonchanie'),
    ('', 'ФИО ответственного', '15', 'fio_otvetstvennogo'),
    ('', 'Мобильный', '16', 'mobilnyy'),
]


HEADER_ALIASES = {
    'округ вуза': 'okrug_vuza',
    'ову отв за подготовку': 'ovu_otv_podgotovku',
    'наименование вуза': 'nazvanie_vuza',
    'вуз': 'nazvanie_vuza',
    'код вус': 'vus_no',
    'вус': 'vus_no',
    'вус наименование': 'vus_naimenovanie',
    'должность': 'doljnost_no',
    'должность наименование': 'doljnost_naimenovanie',
    'сбор стаж': 'sbor_stazhirovka',
    'сбор стажировка': 'sbor_stazhirovka',
    'программа': 'programma_podgotovki',
    'программа подготовки': 'programma_podgotovki',
    'место проведения': 'mesto_provedeniya_uchebnogo_sbora',
    'место проведения сбора': 'mesto_provedeniya_uchebnogo_sbora',
    'преподавателей': 'planiruetsya_prepodavatelej',
    'студентов': 'planiruetsya_studentov',
    'начало': 'srok_provedeniya_nachalo',
    'окончание': 'srok_provedeniya_okonchanie',
    'фио ответственного': 'fio_otvetstvennogo',
    'мобильный': 'mobilnyy',
}


def _normalize(value: str) -> str:
    return ''.join(ch for ch in str(value).lower() if ch.isalnum() or ch.isspace()).strip()


def _normalize_code(value) -> str:
    if value is None:
        return ''

    text = str(value).strip()
    if not text:
        return ''

    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]

    return text


def _code_candidates(value, width: int) -> list[str]:
    text = _normalize_code(value)
    if not text:
        return []

    candidates = [text]
    if text.isdigit():
        padded = text.zfill(width)
        if padded not in candidates:
            candidates.insert(0, padded)
    return candidates


@lru_cache(maxsize=1)
def _load_json_mapping(path: str) -> dict[str, str]:
    file_path = Path(path)
    if not file_path.exists():
        return {}

    try:
        with file_path.open('r', encoding='utf-8') as file_obj:
            raw_mapping = json.load(file_obj)
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(raw_mapping, dict):
        return {}

    normalized: dict[str, str] = {}
    for key, value in raw_mapping.items():
        normalized_key = _normalize_code(key)
        if normalized_key:
            normalized[normalized_key] = str(value)
    return normalized


def _load_decoding_maps() -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    return (
        _load_json_mapping(str(OFFICER_VUS_PATH)),
        _load_json_mapping(str(NONOFFICE_VUS_PATH)),
        _load_json_mapping(str(NONOFFICE_POSITION_PATH)),
    )


def _lookup_decoding(value, mapping: dict[str, str], width: int) -> str | None:
    for candidate in _code_candidates(value, width):
        if candidate in mapping:
            return mapping[candidate]
    return None


def _vus_code_length(value) -> int:
    text = _normalize_code(value)
    digits = ''.join(ch for ch in text if ch.isdigit())
    return len(digits)


def _harmonize_columns(df: pd.DataFrame) -> pd.DataFrame:
    direct_map = {_normalize(column): column for column in REQUIRED_COLUMNS}
    mapped_columns: dict[str, int] = {}

    for col_idx, col in enumerate(df.columns):
        normalized = _normalize(col)
        target = None
        if normalized in direct_map:
            target = direct_map[normalized]
        else:
            for alias, alias_target in HEADER_ALIASES.items():
                if alias in normalized:
                    target = alias_target
                    break

        if target and target not in mapped_columns:
            mapped_columns[target] = col_idx

    out = pd.DataFrame(index=df.index)
    for column in REQUIRED_COLUMNS:
        if column in mapped_columns:
            out[column] = df.iloc[:, mapped_columns[column]].fillna('').astype(str)
        else:
            out[column] = ''

    return out[REQUIRED_COLUMNS]


def _detect_header_depth(raw_df: pd.DataFrame) -> int:
    for idx in range(min(6, len(raw_df))):
        row = raw_df.iloc[idx].astype(str).str.strip()
        numeric_values = [value for value in row if value.isdigit()]
        if len(numeric_values) >= 5 and '1' in numeric_values:
            return idx + 1
    return 1


def _load_tabular_data(file) -> pd.DataFrame:
    if hasattr(file, 'seek'):
        file.seek(0)

    raw_df = pd.read_excel(file, header=None, dtype=str).fillna('')
    header_depth = _detect_header_depth(raw_df)

    if header_depth <= 1:
        if hasattr(file, 'seek'):
            file.seek(0)
        single_header_df = pd.read_excel(file, dtype=str).fillna('')
        return _harmonize_columns(single_header_df)

    header_rows = raw_df.iloc[:header_depth]
    filled_headers = []
    for idx, row in header_rows.iterrows():
        row_values = row.astype(str).str.strip()
        if idx < header_depth - 1:
            row_values = row_values.replace('', pd.NA).ffill().fillna('')
        filled_headers.append(row_values)

    combined_headers = []
    for col_idx in range(raw_df.shape[1]):
        pieces = []
        for row_idx in range(header_depth - 1):
            value = str(filled_headers[row_idx].iloc[col_idx]).strip()
            if value and value not in pieces:
                pieces.append(value)
        combined_headers.append(' '.join(pieces).strip())

    body_df = raw_df.iloc[header_depth:].copy().reset_index(drop=True)
    body_df.columns = combined_headers
    return _harmonize_columns(body_df)


def _export_merged_table(df: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Объединение'

    for col_idx, (group, title, number, column_name) in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=group)
        ws.cell(row=2, column=col_idx, value=title)
        ws.cell(row=3, column=col_idx, value=number)
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = 22

    group_start = None
    current_group = None
    for col_idx, (group, _, _, _) in enumerate(OUTPUT_COLUMNS, start=1):
        if group != current_group:
            if current_group and group_start and col_idx - 1 > group_start:
                ws.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=col_idx - 1)
            current_group = group
            group_start = col_idx
    if current_group and group_start and len(OUTPUT_COLUMNS) > group_start:
        ws.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=len(OUTPUT_COLUMNS))

    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        for col_idx, (_, _, _, column_name) in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[column_name])

    wb.save(path)


def _read_harmonized_dataframe(path_or_file) -> pd.DataFrame:
    return _load_tabular_data(path_or_file)


def _remove_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    cleaned = df.copy().fillna('')
    mask = cleaned.apply(lambda row: any(str(value).strip() for value in row), axis=1)
    return cleaned.loc[mask].reset_index(drop=True)


def merge_excel_files(files) -> Path:
    files = list(files)
    if not files:
        raise ValueError('Не переданы файлы для объединения.')

    frames: list[pd.DataFrame] = []

    for file_obj in files:
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)

        df = _read_harmonized_dataframe(file_obj)
        df = _remove_empty_rows(df)

        if not df.empty:
            frames.append(df)

    if not frames:
        raise ValueError('Не удалось прочитать данные из файлов.')

    merged_df = pd.concat(frames, ignore_index=True)
    merged_df = _remove_empty_rows(merged_df)

    output_dir = Path(settings.MEDIA_ROOT) / 'exports'
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f'merged_{uuid.uuid4().hex}.xlsx'

    _export_merged_table(merged_df, path)
    return path


def decode_for_admin(path: Path) -> Path:
    df = _read_harmonized_dataframe(path)
    officer_vus, nonoffice_vus, nonoffice_positions = _load_decoding_maps()

    def _decode_vus(row) -> str:
        if _vus_code_length(row['vus_no']) == 6:
            officer_name = _lookup_decoding(row['vus_no'], officer_vus, 6)
            if officer_name:
                return officer_name
        elif _vus_code_length(row['vus_no']) == 3:
            nonoffice_name = _lookup_decoding(row['vus_no'], nonoffice_vus, 3)
            if nonoffice_name:
                return nonoffice_name

        return row['vus_naimenovanie']

    df['vus_naimenovanie'] = df.apply(_decode_vus, axis=1)

    officer_mask = df['vus_no'].map(lambda value: _vus_code_length(value) == 6)
    nonofficer_mask = df['vus_no'].map(lambda value: _vus_code_length(value) == 3)

    df.loc[officer_mask, ['doljnost_no', 'doljnost_naimenovanie']] = ''
    df.loc[nonofficer_mask, 'doljnost_naimenovanie'] = df.loc[nonofficer_mask, 'doljnost_no'].map(
        lambda val: _lookup_decoding(val, nonoffice_positions, 3) or ''
    )

    decoded_path = path.with_name(f'{path.stem}_decoded.xlsx')
    _export_merged_table(df, decoded_path)
    return decoded_path


def _apply_border_to_range(ws, r1, c1, r2, c2, border):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def _fill_row(ws, row, c1, c2, fill, border):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = border


def _okrug_full_name(abbr: str) -> str:
    if not isinstance(abbr, str):
        return 'Военный округ'
    key = abbr.strip().upper().replace('Ё', 'Е')
    mapping = {
        'МВО': 'Московский военный округ',
        'ЛЕНВО': 'Ленинградский военный округ',
        'ЮВО': 'Южный военный округ',
        'ЦВО': 'Центральный военный округ',
        'ВВО': 'Восточный военный округ',
    }
    return mapping.get(key, f'{abbr} военный округ')


def _write_section(wb: Workbook, title: str, df: pd.DataFrame, is_sergeants: bool):
    fill_title = PatternFill('solid', fgColor='C6E0B4')
    fill_header = PatternFill('solid', fgColor='FFE699')
    fill_okrug = PatternFill('solid', fgColor='F8CBAD')
    fill_ovu = PatternFill('solid', fgColor='BDD7EE')
    fill_sub = PatternFill('solid', fgColor='C6E0B4')
    fill_grand = PatternFill('solid', fgColor='7030A0')
    fill_a_orange = PatternFill('solid', fgColor='C6E0B4')

    side = Side(border_style='medium', color='000000')
    border = Border(left=side, right=side, top=side, bottom=side)

    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

    if not is_sergeants:
        max_col = 8
        col_stud = 3
        col_teach = 7
        detail_cols = [
            'vus_no', 'vus_naimenovanie', 'planiruetsya_studentov',
            'mesto_provedeniya_uchebnogo_sbora',
            'srok_provedeniya_nachalo', 'srok_provedeniya_okonchanie',
            'planiruetsya_prepodavatelej', 'sbor_stazhirovka',
        ]
    else:
        max_col = 9
        col_stud = 5
        col_teach = 9
        detail_cols = [
            'vus_no', 'vus_naimenovanie',
            'doljnost_no', 'doljnost_naimenovanie',
            'planiruetsya_studentov',
            'mesto_provedeniya_uchebnogo_sbora',
            'srok_provedeniya_nachalo', 'srok_provedeniya_okonchanie',
            'planiruetsya_prepodavatelej',
        ]

    ws = wb.create_sheet(title=title)
    ws.default_row_height = 15
    r = 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    c = ws.cell(row=r, column=1, value=title)
    c.font, c.alignment = font_title, center_wrap
    _apply_border_to_range(ws, r, 1, r, max_col, border)
    r += 1

    if not is_sergeants:
        top = ['Код ВУС', 'Наименование ВУС', 'Кол-во студентов', 'Место сбора',
               'Срок проведения', None, 'Кол-во препод.', 'Сбор/стажировка']
        bot = [None, None, None, None, 'начало', 'окончание', None, None]
        span_c1 = 5
    else:
        top = ['№ ВУС', 'Наименование ВУС', '№ должности', 'Наименование должности',
               'Кол-во студентов', 'Место сбора', 'Срок проведения', None, 'Кол-во препод.']
        bot = [None, None, None, None, None, None, 'начало', 'окончание', None]
        span_c1 = 7

    for col, text in enumerate(top, start=1):
        if text is None:
            continue
        if text == 'Срок проведения':
            ws.merge_cells(start_row=r, start_column=span_c1, end_row=r, end_column=span_c1 + 1)
            cc = ws.cell(row=r, column=span_c1, value=text)
            cc.font, cc.alignment = font_bold, center_wrap
            _apply_border_to_range(ws, r, span_c1, r, span_c1 + 1, border)
        else:
            ws.merge_cells(start_row=r, start_column=col, end_row=r + 1, end_column=col)
            cc = ws.cell(row=r, column=col, value=text)
            cc.font, cc.alignment = font_bold, center_wrap
            _apply_border_to_range(ws, r, col, r + 1, col, border)

    for col, text in enumerate(bot, start=1):
        if text:
            cc = ws.cell(row=r + 1, column=col, value=text)
            cc.font, cc.alignment, cc.border = font_bold, center_wrap, border
    r += 2

    grand_stud = grand_teach = 0
    for okrug, g_ok in df.groupby('okrug_vuza', sort=False):
        okrug_abbr = str(okrug or '').strip().upper()
        okrug_full = _okrug_full_name(okrug_abbr)

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
        c = ws.cell(row=r, column=1, value=okrug_full)
        c.font, c.fill, c.alignment = font_bold, fill_okrug, center_wrap
        _apply_border_to_range(ws, r, 1, r, max_col, border)
        r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
        c = ws.cell(
            row=r,
            column=1,
            value='На территории военного округа по месту расположения образовательной организации',
        )
        c.font, c.fill, c.alignment = font_bold, fill_title, center_wrap
        _apply_border_to_range(ws, r, 1, r, max_col, border)
        r += 1

        ok_stud = int(g_ok['planiruetsya_studentov'].sum())
        ok_teach = int(g_ok['planiruetsya_prepodavatelej'].sum())

        for ovu, g_ovu in g_ok.groupby('ovu_otv_podgotovku', sort=False):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
            c = ws.cell(row=r, column=1, value=str(ovu))
            c.font, c.fill, c.alignment = font_bold, fill_ovu, center_wrap
            _apply_border_to_range(ws, r, 1, r, max_col, border)
            r += 1

            for vuza, g_vuz in g_ovu.groupby('nazvanie_vuza', sort=False):
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
                c = ws.cell(row=r, column=1, value=str(vuza))
                c.font, c.fill, c.alignment = font_bold, fill_title, center_wrap
                _apply_border_to_range(ws, r, 1, r, max_col, border)
                r += 1

                for _, rec in g_vuz.iterrows():
                    vals = []
                    for key in detail_cols:
                        val = rec.get(key, '')
                        if 'srok_provedeniya' in key and pd.notna(val):
                            val = pd.to_datetime(val).strftime('%d.%m.%Y')
                        vals.append(val)
                    for i, value in enumerate(vals, start=1):
                        cell = ws.cell(row=r, column=i, value=value)
                        cell.alignment = center_wrap
                        cell.border = border
                    r += 1

            sub_stud = int(g_ovu['planiruetsya_studentov'].sum())
            sub_teach = int(g_ovu['planiruetsya_prepodavatelej'].sum())

            ws.cell(row=r, column=1).fill = fill_a_orange
            ws.cell(row=r, column=1).border = border
            _fill_row(ws, r, 2, max_col, fill_sub, border)

            lab = ws.cell(row=r, column=2, value=f'Всего за {ovu}')
            lab.font, lab.alignment = font_bold, center_wrap

            s = ws.cell(row=r, column=col_stud, value=sub_stud)
            s.fill, s.border, s.alignment = fill_header, border, center_wrap
            t = ws.cell(row=r, column=col_teach, value=sub_teach)
            t.fill, t.border, t.alignment = fill_header, border, center_wrap
            r += 1

        ws.cell(row=r, column=1).fill = fill_grand
        ws.cell(row=r, column=1).border = border
        _fill_row(ws, r, 2, max_col, fill_grand, border)

        lab = ws.cell(row=r, column=2, value=f'ВСЕГО ЗА {okrug_abbr}')
        lab.font, lab.alignment = font_bold, center_wrap

        s = ws.cell(row=r, column=col_stud, value=ok_stud)
        s.fill, s.border, s.alignment = fill_header, border, center_wrap
        t = ws.cell(row=r, column=col_teach, value=ok_teach)
        t.fill, t.border, t.alignment = fill_header, border, center_wrap
        r += 1

        grand_stud += ok_stud
        grand_teach += ok_teach

    ws.cell(row=r, column=1).fill = fill_grand
    ws.cell(row=r, column=1).border = border
    _fill_row(ws, r, 2, max_col, fill_grand, border)

    lab = ws.cell(row=r, column=2, value='ИТОГО')
    lab.font, lab.alignment = Font(bold=True, size=12), center_wrap

    s = ws.cell(row=r, column=col_stud, value=grand_stud)
    s.fill, s.border, s.alignment = fill_header, border, center_wrap
    t = ws.cell(row=r, column=col_teach, value=grand_teach)
    t.fill, t.border, t.alignment = fill_header, border, center_wrap

    ws.column_dimensions[get_column_letter(2)].width = 50


def create_report(path: Path) -> Path:
    df = _read_harmonized_dataframe(path)
    for col in ('srok_provedeniya_nachalo', 'srok_provedeniya_okonchanie'):
        df[col] = pd.to_datetime(df[col], errors='coerce')
    for col in ('planiruetsya_studentov', 'planiruetsya_prepodavatelej'):
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    cadre_df = df[df['programma_podgotovki'] == 'Офицеры кадра']
    reserve_df = df[df['programma_podgotovki'] == 'Офицеры запаса']
    others_df = df[~df['programma_podgotovki'].isin(['Офицеры кадра', 'Офицеры запаса'])]

    wb = Workbook()
    wb.remove(wb.active)
    _write_section(wb, 'I. Офицеры кадра', cadre_df, is_sergeants=False)
    _write_section(wb, 'II. Офицеры запаса', reserve_df, is_sergeants=False)
    _write_section(wb, 'III. Сержанты и солдаты', others_df, is_sergeants=True)

    report_path = path.with_name(f'{path.stem}_report.xlsx')
    wb.save(report_path)
    return report_path


def load_editor_payload(path: Path) -> tuple[list[str], list[list[str]]]:
    df = _read_harmonized_dataframe(path)
    df = _remove_empty_rows(df)

    column_names = [column_name for _, _, _, column_name in OUTPUT_COLUMNS]
    column_titles = [title for _, title, _, _ in OUTPUT_COLUMNS]

    if df.empty:
        return column_titles, []

    rows = df[column_names].fillna('').astype(str).values.tolist()
    return column_titles, rows


def save_editor_rows(path: Path, rows: list[list[str]]) -> None:
    column_names = [column_name for _, _, _, column_name in OUTPUT_COLUMNS]

    normalized_rows: list[list[str]] = []
    for row in rows:
        if not isinstance(row, list):
            continue
        clipped = [str(cell) if cell is not None else '' for cell in row[:len(column_names)]]
        if len(clipped) < len(column_names):
            clipped.extend([''] * (len(column_names) - len(clipped)))
        normalized_rows.append(clipped)

    df = pd.DataFrame(normalized_rows, columns=column_names)
    df = _remove_empty_rows(df)
    _export_merged_table(df, path)
